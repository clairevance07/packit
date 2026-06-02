import streamlit as st
import temp_data
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from io import BytesIO

def make_xlsx(formatted_liquids, formatted_clothes, formatted_toiletries):
    wb = Workbook()
    ws = wb.active

    border_thickness = Side(border_style='thin', color='000000')
    cell_border = Border(top=border_thickness, bottom=border_thickness, left=border_thickness, right=border_thickness)

    ws['A1'] = 'Liquids'
    ws['A1'].font = Font(bold=True)
    max_a_len = 0
    for i in range(len(formatted_liquids)):
        ws[f'A{i+2}'] = formatted_liquids[i]
        ws[f'A{i+2}'].border = cell_border
        if len(formatted_liquids[i]) > max_a_len:
            max_a_len = len(formatted_liquids[i])
    ws.column_dimensions['A'].width = max_a_len

    ws['C1'] = 'Toiletries'
    ws['C1'].font = Font(bold=True)
    for i in range(len(formatted_toiletries)):
        ws[f'C{i+2}'] = formatted_toiletries[i]
        ws[f'C{i+2}'].border = cell_border
        if len(formatted_toiletries[i]) > max_a_len:
            max_a_len = len(formatted_toiletries[i])
    ws.column_dimensions['C'].width = max_a_len

    ws['E1'] = 'Clothing'
    ws['E1'].font = Font(bold=True)
    for i in range(len(formatted_clothes)):
        ws[f'E{i+2}'] = formatted_clothes[i]
        ws[f'E{i+2}'].border = cell_border
        if len(formatted_clothes[i]) > max_a_len:
            max_a_len = len(formatted_clothes[i])
    ws.column_dimensions['E'].width = max_a_len

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

temperature = None

st.title("PackIt")
st.subheader("Welcome! Let's help you make a packing list for your stateside vacation!")
st.divider()

destination = st.selectbox("What is your travel destination?", [" ", "Alabama (AL)", "Alaska (AK)", "Arizona (AZ)", "Arkansas (AR)", "California (CA)", "Colorado (CO)", "Connecticut (CT)", "Delaware (DE)", "Florida (FL)", "Georgia (GA)", "Hawaii (HI)", "Idaho (ID)", "Illinois (IL)", "Indiana (IN)", "Iowa (IA)", "Kansas (KS)", "Kentucky (KY)", "Louisiana (LA)", "Maine (ME)", "Maryland (MD)", "Massachusetts (MA)", "Michigan (MI)", "Minnesota (MN)", "Mississippi (MS)", "Missouri (MO)", "Montana (MT)", "Nebraska (NE)", "Nevada (NV)", "New Hampshire (NH)", "New Jersey (NJ)", "New Mexico (NM)", "New York (NY)", "North Carolina (NC)", "North Dakota (ND)", "Ohio (OH)", "Oklahoma (OK)", "Oregon (OR)", "Pennsylvania (PA)", "Rhode Island (RI)", "South Carolina (SC)", "South Dakota (SD)", "Tennessee (TN)", "Texas (TX)", "Utah (UT)", "Vermont (VT)", "Virginia (VA)", "Washington (WA)", "West Virginia (WV)", "Wisconsin (WI)", "Wyoming (WY)"])
st.divider()

temperature_set = temp_data.find_destination(destination)

month = st.selectbox(f"What month are you traveling?", [" ", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
st.divider()

if temperature_set is not None and month is not None:
    temperature = temp_data.get_temp(month, temperature_set)

days = st.selectbox("How long is your vacation?", [" ", "1-3 days", "4-7 days", "8+ days"])
st.divider()

liquids = st.multiselect("Please select the liquids you will need for your trip: ", ["Liquid makeup", "Contact solution", "Extra contacts", "Face wash", "Toner", "Face lotion", "Mouthwash", "Toothpaste", "Sunscreen", "Body lotion", "Eye drops", "Perfume", "Makeup remover", "Essential hair products", "Shaving cream", "Detergent"])
st.divider()

toiletries = st.multiselect("Please select the additional toiletries you will need for your trip:", ["Tweezers", "Medication", "Contact case", "Makeup", "Cotton rounds", "Hair accessories", "Jewelry", "Toothbrush", "Floss", "Razor", "Feminine products", "Glasses", "Retainer", "Brush/comb", "Deodorant", "Lip balm", "Cotton swabs", "Nail file/clipper"], default=["Brush/comb", "Toothbrush", "Floss", "Razor", "Deodorant", "Lip balm", "Cotton swabs", "Nail file/clipper"])
st.divider()

def calc_warm(vacay_length):
    if vacay_length == "1-3 days":
        return {
        "short_sleeved": 3,
        "sweatshirt": 1,
        "pair_shorts": 2,
        "pair_pants": 1,
        "pair_pajamas": 1,
        "pair_undergarments": "1-3",
        "pair_socks": "1-3",
        "pair_athletic": 2,
        "formal": 1,
        "shoes": 2
    }

    if vacay_length == "4-7 days":
        return {
        "short_sleeved": 5,
        "sweatshirt": 2,
        "pair_shorts": 3,
        "pair_pants": 1,
        "pair_pajamas": "4-7",
        "pair_undergarments": "4-7",
        "pair_socks": "4-7",
        "pair_athletic": 2,
        "formal": 1,
        "shoes": 3
    }

    if vacay_length == "8+ days":
        return {
        "short_sleeved": 7,
        "sweatshirt": 3,
        "pair_shorts": 4,
        "pair_pants": 2,
        "pair_pajamas": 2,
        "pair_undergarments": "8+",
        "pair_socks": "8+",
        "pair_athletic": 2,
        "formal": 1,
        "shoes": 3
    }
    
def calc_mod(vacay_length):
    if vacay_length == "1-3 days":
        return {
        "short_sleeved": 2,
        "long_sleeved": 1,
        "jacket": 1,
        "sweatshirt": 1,
        "coat": 1,
        "pair_pants": 2,
        "pair_pajamas": 1,
        "pair_undergarments": "1-3",
        "pair_socks": "1-3",
        "pair_athletic": 2,
        "formal": 1,
        "shoes": 2
    }

    if vacay_length == "4-7 days":
        return {
        "short_sleeved": 3,
        "long_sleeved": 2,
        "jacket": 1,
        "sweatshirt": 2,
        "coat": 1,
        "pair_pants": 4,
        "pair_pajamas": 2,
        "pair_undergarments": "4-7",
        "pair_socks": "4-7",
        "pair_athletic": 2,
        "formal": 1,
        "shoes": 3
    }

    if vacay_length == "8+ days":
        return {
        "short_sleeved": 4,
        "long_sleeved": 3,
        "jacket": 1,
        "sweatshirt": 2,
        "coat": 1,
        "pair_pants": 6,
        "pair_pajamas": 2,
        "pair_undergarments": "8+",
        "pair_socks": "8+",
        "pair_athletic": 2,
        "formal": 1,
        "shoes": 3
    }
        
def calc_cold(vacay_length):
    if vacay_length == "1-3 days":
        return {
        "long_sleeved": 3,
        "sweatshirt": 1, 
        "coat": 1,
        "pair_pants": 2,
        "pair_pajamas": 1,
        "pair_undergarments": "1-3",
        "pair_socks": "1-3",
        "pair_athletic": 2,
        "formal": 1,
        "hat": 1,
        "pair_gloves": 1,
        "shoes": 2
    }

    if vacay_length == "4-7 days":
        return {
        "long_sleeved": 5,
        "sweatshirt": 2, 
        "coat": 1,
        "pair_pants": 4,
        "pair_pajamas": 2,
        "pair_undergarments": "4-7",
        "pair_socks": "4-7",
        "pair_athletic": 2,
        "formal": 1,
        "hat": 1,
        "pair_gloves": 1,
        "shoes": 3
    }

    if vacay_length == "8+ days":
        return {
        "long_sleeved": 7,
        "sweatshirt": 2, 
        "coat": 1,
        "pair_pants": 6,
        "pair_pajamas": 2,
        "pair_undergarments": "8+",
        "pair_socks": "8+",
        "pair_athletic": 2,
        "formal": 1,
        "hat": 1,
        "pair_gloves": 1,
        "shoes": 3
    }

if st.button("Generate Packing List"):
    if temperature is not None and days != " " and liquids != [] and toiletries != []:
        bar = st.progress(0, text="Generating Packing List")
        for second in range(100):
            time.sleep(0.01)
            bar.progress(second + 1, text="Generating Packing List")
        time.sleep(1)
        bar.empty()
        st.divider()
        if temperature >= 70:
            clothes = calc_warm(days)
        elif temperature >= 55:
            clothes = calc_mod(days)
        else:
            clothes = calc_cold(days)
        st.header(f"Packing list for {month} trip to {destination.split("(")[0]}")
        st.subheader("Clothing:")

        formatted_clothes = []
        for item in clothes:
            if len(item.split("_")) > 1:
                if item.split("_")[1] == "sleeved":
                    st.write(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]} shirts")
                    formatted_clothes.append(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]} shirts")
                elif item.split("_")[1] == "athletic":
                    st.write(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]} clothes")
                    formatted_clothes.append(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]} clothes")
                elif item.split("_")[1] == "undergarments" or item.split("_")[1] == "socks":
                    st.write(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]} (one pair for each day of your trip)")
                    formatted_clothes.append(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]} (one pair for each day of your trip)")
                else:
                    st.write(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]}")
                    formatted_clothes.append(f"{clothes[item]} {item.split("_")[0]} {item.split("_")[1]}")
            elif clothes[item] > 1 and item[-1] != 's':
                st.write(f"{clothes[item]} {item}s")
                formatted_clothes.append(f"{clothes[item]} {item}s")
            elif item == "formal":
                st.write(f"{clothes[item]} formal / church outfit")
                formatted_clothes.append(f"{clothes[item]} formal / church outfit")
            elif item == "shoes":
                if clothes[item] == 2:
                    st.write("1 pair tennis shoes")
                    st.write("1 pair nicer shoes")
                    formatted_clothes.append("1 pair tennis shoes")
                    formatted_clothes.append("1 pair nicer shoes")
                else:
                    st.write("1 pair tennis shoes")
                    st.write("1 pair sneakers")
                    st.write("1 pair nicer shoes")
                    formatted_clothes.append("1 pair tennis shoes")
                    formatted_clothes.append("1 pair sneakers")
                    formatted_clothes.append("1 pair nicer shoes")
            else:
                st.write(f"{clothes[item]} {item}")
                formatted_clothes.append(f"{clothes[item]} {item}")
        
        st.subheader("Liquids:")
        st.caption("Each liquid must be in a bottle that is 3.4 oz or less. All liquids must fit in one clear, quart-sized bag.")

        formatted_liquids = []
        for item in liquids:
            st.write(item)
            formatted_liquids.append(item)

        formatted_toiletries = []
        st.subheader("Toiletries:")
        for item in toiletries:
            st.write(item)
            formatted_toiletries.append(item)

        st.space()

        excel_file = make_xlsx(formatted_liquids, formatted_clothes, formatted_toiletries)
        st.download_button(
            "Download Packing List Spreadsheet",
            data=excel_file.getvalue(),
            file_name=f"{destination}_packing_list_{month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("You have not selected a value in all fields.")